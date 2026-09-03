import streamlit as st
import pandas as pd
import io
import json
import os
import google.generativeai as genai
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# 1. UI Configuration & API
st.set_page_config(page_title="BEMS Estimator - DC Controls", layout="wide")
st.title("Automated BEMS Points List & Estimator")
st.markdown("Generador basado en el estándar de cotización de DC Controls (Formato IDA Cavan).")

api_key = st.text_input("Enter API Key (Gemini/OpenAI):", type="password")
if api_key:
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel('gemini-1.5-flash') 

# 2. Knowledge Base (Real rules extracted from IDA Cavan Project)
engineering_rules = {
    "Common LPHW/CHW Devices (System Level)": {
        "mandatory": [
            "Header Flow Immersion Temperature Sensor",
            "Header Return Immersion Temperature Sensor",
            "Outside Frost Thermostat",
            "Outside Temperature Sensor",
            "Immersion Frost Thermostat"
        ],
        "notes": "Include these ONLY ONCE per plant for the main headers."
    },
    "Boiler": {
        "per_unit": [
            "Boiler Enable",
            "Boiler Common Fault",
            "Boiler Control Signal",
            "Boiler Flow Immersion Temperature Sensor",
            "Boiler Return Immersion Temperature Sensor"
        ]
    },
    "Pump (Primary/Secondary)": {
        "per_unit": [
            "Pump Enable",
            "Pump Status",
            "Variable Speed Drive",
            "Flow Immersion Temperature Sensor",
            "3 Port Control Valve / Actuator"
        ]
    },
    "Pressurisation Unit": {
        "per_unit": [
            "Pressurisation Unit High Pressure",
            "Pressurisation Unit Low Pressure"
        ]
    },
    "Calorifier / Hot Water Generator": {
        "per_unit": [
            "Enable",
            "Common Fault",
            "Control Signal",
            "Immersion Temperature Sensor",
            "High Limit Thermostat (60-95 Man. Reset)"
        ]
    },
    "AHU (Air Handling Unit)": {
        "per_unit": [
            "Enable", "Status", "Control Signal", "Supply Air Temp Sensor", "Return Air Temp Sensor", "Frost Stat"
        ]
    },
    "FCU (Fan Coil Unit)": {
        "per_unit": [
            "Space Temperature Sensor", "Control Valve Actuator"
        ]
    }
}

# 3. Technical Catalog (Exact extraction from IDA Cavan)
# Base Labour is 50 per point. The AI will multiply this by the quantity.
component_catalog = {
    "Header Flow Immersion Temperature Sensor": {"Part": "TTI-S Brass Pocket", "AI": 1, "AO": 0, "DI": 0, "DO": 0, "Labour": 50},
    "Header Return Immersion Temperature Sensor": {"Part": "TTI-S Brass Pocket", "AI": 1, "AO": 0, "DI": 0, "DO": 0, "Labour": 50},
    "Outside Frost Thermostat": {"Part": "DBET-23U", "AI": 0, "AO": 0, "DI": 1, "DO": 0, "Labour": 50},
    "Outside Temperature Sensor": {"Part": "TB/TO", "AI": 1, "AO": 0, "DI": 0, "DO": 0, "Labour": 50},
    "Immersion Frost Thermostat": {"Part": "DBTV-2U", "AI": 0, "AO": 0, "DI": 1, "DO": 0, "Labour": 50},
    
    "Boiler Enable": {"Part": "Volt Free Contacts", "AI": 0, "AO": 0, "DI": 0, "DO": 1, "Labour": 50},
    "Boiler Common Fault": {"Part": "Volt Free Contacts", "AI": 0, "AO": 0, "DI": 1, "DO": 0, "Labour": 50},
    "Boiler Control Signal": {"Part": "0…10V dc", "AI": 0, "AO": 1, "DI": 0, "DO": 0, "Labour": 50},
    "Boiler Flow Immersion Temperature Sensor": {"Part": "TTI-S Brass Pocket", "AI": 1, "AO": 0, "DI": 0, "DO": 0, "Labour": 50},
    "Boiler Return Immersion Temperature Sensor": {"Part": "TTI-S Brass Pocket", "AI": 1, "AO": 0, "DI": 0, "DO": 0, "Labour": 50},
    
    "Pump Enable": {"Part": "Volt Free Contacts", "AI": 0, "AO": 0, "DI": 0, "DO": 1, "Labour": 50},
    "Pump Status": {"Part": "Volt Free Contacts", "AI": 0, "AO": 0, "DI": 1, "DO": 0, "Labour": 50},
    "Variable Speed Drive": {"Part": "Built in to Pump", "AI": 0, "AO": 1, "DI": 0, "DO": 0, "Labour": 50},
    "Flow Immersion Temperature Sensor": {"Part": "TTI-S Brass Pocket", "AI": 1, "AO": 0, "DI": 0, "DO": 0, "Labour": 50},
    "3 Port Control Valve / Actuator": {"Part": "Valve 40mm", "AI": 0, "AO": 1, "DI": 0, "DO": 0, "Labour": 50},
    
    "Pressurisation Unit High Pressure": {"Part": "Volt Free Contacts", "AI": 0, "AO": 0, "DI": 1, "DO": 0, "Labour": 50},
    "Pressurisation Unit Low Pressure": {"Part": "Volt Free Contacts", "AI": 0, "AO": 0, "DI": 1, "DO": 0, "Labour": 50},
    
    "Enable": {"Part": "Volt Free Contacts", "AI": 0, "AO": 0, "DI": 0, "DO": 1, "Labour": 50},
    "Common Fault": {"Part": "Volt Free Contacts", "AI": 0, "AO": 0, "DI": 1, "DO": 0, "Labour": 50},
    "Status": {"Part": "Volt Free Contacts", "AI": 0, "AO": 0, "DI": 1, "DO": 0, "Labour": 50},
    "Control Signal": {"Part": "0…10V dc", "AI": 0, "AO": 1, "DI": 0, "DO": 0, "Labour": 50},
    "Immersion Temperature Sensor": {"Part": "TI/Brass Pocket", "AI": 1, "AO": 0, "DI": 0, "DO": 0, "Labour": 50},
    "High Limit Thermostat (60-95 Man. Reset)": {"Part": "RAK TW 1000B", "AI": 0, "AO": 0, "DI": 1, "DO": 0, "Labour": 50},
    
    "Space Temperature Sensor": {"Part": "RS-Temp", "AI": 1, "AO": 0, "DI": 0, "DO": 0, "Labour": 50},
    "Control Valve Actuator": {"Part": "MVC / DB_VZ", "AI": 0, "AO": 1, "DI": 0, "DO": 0, "Labour": 50},
    "Supply Air Temp Sensor": {"Part": "Duct Temp Sensor", "AI": 1, "AO": 0, "DI": 0, "DO": 0, "Labour": 50},
    "Return Air Temp Sensor": {"Part": "Duct Temp Sensor", "AI": 1, "AO": 0, "DI": 0, "DO": 0, "Labour": 50},
    "Frost Stat": {"Part": "DBET-23U", "AI": 0, "AO": 0, "DI": 1, "DO": 0, "Labour": 50}
}

project_description = st.text_area(
    "Describe the project for the Points List:", 
    placeholder="Example: We need a plant with 2 Boilers, 3 Primary Pumps, 2 LPHW Pressurisation Units, and 1 Calorifier."
)

# 4. Generation Logic
if st.button("Generate Points List (Excel)"):
    if not api_key or not project_description:
        st.warning("Please provide both the API Key and the project description.")
    else:
        with st.spinner("Engineering Points List, assigning real IOs, Part Numbers and calculating Labour..."):
            
            prompt = f"""
            You are an expert BEMS estimator working for DC Controls. Generate a Points List based on the description, mimicking the exact style of the "IDA Cavan" project.
            
            ENGINEERING RULES (What components go in each system):
            {json.dumps(engineering_rules, indent=2)}
            
            TECHNICAL CATALOG (Exact Parts and I/O values):
            {json.dumps(component_catalog, indent=2)}
            
            CRITICAL FORMATTING INSTRUCTIONS (CAVAN PROJECT STYLE):
            1. Create a HEADER ROW for each main equipment group. (e.g., Description: "Boiler", Quantity: 2, MCC: "MCB". Leave AI, AO, DI, DO, Labour blank).
            2. Below the header row, list its components based on the ENGINEERING RULES.
            3. CRITICAL CALCULATION: For each component, lookup its base AI, AO, DI, DO, and Labour in the CATALOG. You MUST multiply these base values by the quantity of the main equipment.
            4. Use the exact Part No. from the catalog (e.g., "Volt Free Contacts", "0...10V dc", "TTI-S Brass Pocket").
            5. Leave IOs or Labour as empty strings ("") if the value is 0.
            
            User description: "{project_description}"
            
            Return ONLY a JSON array matching the standard DC Controls columns. DO NOT use markdown.
            [
              {{
                "Description": "Boiler", "AI": "", "AO": "", "DI": "", "DO": "", "MCC": "MCB", "Quantity": 2, "Part No.": "", "Panel At 20%": "", "Parts At 0%": "", "Labour At 20%": ""
              }},
              {{
                "Description": "Boiler Flow Immersion Temperature Sensor", "AI": 2, "AO": "", "DI": "", "DO": "", "MCC": "", "Quantity": 2, "Part No.": "TTI-S Brass Pocket", "Panel At 20%": "", "Parts At 0%": "", "Labour At 20%": 100
              }}
            ]
            """
            
            try:
                # Llamada a la IA
                response = model.generate_content(prompt)
                json_text = response.text.strip().replace("```json", "").replace("```", "")
                materials_data = json.loads(json_text)
                
                # --- AQUÍ ESTÁ LA MAGIA DEL FORMATO OFICIAL ---
                template_path = "template.xlsx"
                
                if not os.path.exists(template_path):
                    st.error("⚠️ Falta el archivo 'template.xlsx' en tu GitHub. Súbelo para usar el formato oficial.")
                else:
                    # Abrir la plantilla sin romper su diseño
                    wb = load_workbook(template_path)
                    ws = wb["Points List"] # Asegúrate de que la hoja se llame así en tu template
                    
                    # El renglón donde empiezan los datos reales (según tu Excel, debajo de Quote Ref, Project, etc.)
                    # Usualmente es el renglón 6 o 7. Ajusta este número si tus datos empiezan más abajo.
                    start_row = 6 
                    
                    for idx, row_data in enumerate(materials_data):
                        current_row = start_row + idx
                        # Insertamos columna por columna basándonos en tu layout:
                        # Columna B = Description, C = AI, D = AO, etc.
                        ws.cell(row=current_row, column=2, value=row_data.get("Description", ""))
                        ws.cell(row=current_row, column=3, value=row_data.get("AI", ""))
                        ws.cell(row=current_row, column=4, value=row_data.get("AO", ""))
                        ws.cell(row=current_row, column=5, value=row_data.get("DI", ""))
                        ws.cell(row=current_row, column=6, value=row_data.get("DO", ""))
                        ws.cell(row=current_row, column=7, value=row_data.get("MCC", ""))
                        ws.cell(row=current_row, column=8, value=row_data.get("Quantity", ""))
                        ws.cell(row=current_row, column=9, value=row_data.get("Part No.", ""))
                        ws.cell(row=current_row, column=10, value=row_data.get("Panel At 20%", ""))
                        ws.cell(row=current_row, column=11, value=row_data.get("Parts At 0%", ""))
                        ws.cell(row=current_row, column=12, value=row_data.get("Labour At 20%", ""))

                    buffer = io.BytesIO()
                    wb.save(buffer)

                    st.success("¡Cotización generada usando el formato oficial de DC Controls!")
                    st.download_button(
                        label="📥 Descargar Points List Oficial (Excel)",
                        data=buffer.getvalue(),
                        file_name="DC_Controls_Quotation.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    # Mostrar una tablita rápida en la web (opcional, solo visual)
                    st.dataframe(pd.DataFrame(materials_data), use_container_width=True)

            except Exception as e:
                st.error(f"Error: {e}\nRevisa que el JSON de la IA sea válido.")
